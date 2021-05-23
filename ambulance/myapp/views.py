from django.shortcuts import render
import requests
import re
import pandas as pd
import openpyxl
import smtplib
import os

# Create your views here.
name=""
number=""
address=""
email=""
city=""
def home(request):
    return render(request,"home.html")


def second(request):
    global name,number,address,email,city,opnumber
    name=request.GET.get("name")
    number=request.GET.get("number")
    address=request.GET.get("address")
    city=request.GET.get("city")
    email=request.GET.get("cars")
    opnumber=request.GET.get("op")


    new_address=address+" "+city    
        
    succ={}
    ads=""
    
    for a in new_address.split():
        ads+=a
        ads+='+'
    

    link="https://maps.google.com/maps?q=hospital+near+"+ads[0:-1]+"&output=embed&z=13"
    succ={"num":link}
    print("i am harsh >>>>>",name,number,address,email,link)
    

    
    
    return render(request,"second.html",succ)



def select_hospital(request):
    global message
    hospital=request.GET.get("hospital")
    message=request.GET.get("message")
    site1=[]
    succ={}
    double_check=""
    succ["name"]=name
    succ["number"]=number
    succ["address"]=address
    succ["email"]=email
    succ["message"]=message
    
    

#hospital searching from database
    hospital_name=hospital
    hospital_name1=hospital_name.lower()
    

    n=1
    empty=""
    file=pd.ExcelFile("database.xlsx")
    sheet=file.parse("Sheet1")
    for x in sheet["Hospital Name "]:
        y=x.lower()        
        n+=1
        if y.find(hospital_name1)==-1:
            continue        
        else:
            
            empty="found"
            break
        

    if empty=="found":
        wb = openpyxl.load_workbook("database.xlsx")
          
        sheet = wb.active

        global num1,num2,num3,num4,num5
           
        x1 = sheet.cell(row=n ,column=1)  
        x2 = sheet.cell(row=n ,column=2)  
        x3 = sheet.cell(row=n ,column=3)  
        x4 = sheet.cell(row=n ,column=4)  
        x5 = sheet.cell(row=n ,column=5)

        num1=x1.value
        num2=x2.value
        num3=x3.value
        num4=x4.value
        num5=x5.value
        
        succ["email"]=email

        succ["Hospitalname"]=x1.value
        succ["Hospitalemail"]=x2.value
        succ["Hospitalphonenumber"]=x3.value
        succ["Hospitaladdress"]=x4.value
        succ["Hospitalpincode"]=x5.value       
        
    else:
        double_check="Not found"

    if double_check=="Not found":
        succ["nohospital"]=hospital
        return render(request,"math.html",succ)      
              
        
    else:
        return render(request,"request.html",succ)
        
    
def send_mail(request):
    print()
    
    Email_address=os.environ.get('USER_NAME')
    Password=os.environ.get('PASSWORD')

    with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
        smtp.ehlo()
        smtp.starttls()
        smtp.ehlo()
        

        #hospital detals
        NAME=num1
        EMAIL=num2
        PHONE_NO=num3
        ADDRESS=num4
        PINCODE=num5
        succ={'hospital':NAME,'opnumber':opnumber,"hoscontact":PHONE_NO}
        

        if EMAIL==None:
            succ["hospitalno"]=PHONE_NO            
            return render(request,"email_not_found.html",succ)

        else:

            smtp.login(Email_address, Password)
            subject="EMERGENCY REQUEST"
            body='REQUEST TO SEND THE AMBULANCE TO THE GIVEN DETAILS\n\n\n\n\nUSER DETAILS\n\nNAME : '+name.upper()+"\n"+'ADDRESS :'+address+"\n"+'PHONE NO : '+number+"\n"+"OPTIONAL PH. No : "+opnumber+"\n"+'EMAIL : '+email+"\n"+'MESSAGE : '+message+"\n"+"\n\nHOSPITAL DETAILS\n\nNAME :"+NAME+"\n"+'EMAIL : '+EMAIL+"\n"+'PHONE NUMBER : '+PHONE_NO+"\n"+'ADDRESS : '+ADDRESS+""+city+"\n"+"PIN CODE :"+str(PINCODE)

            msg=f'Subject:{subject}\n\n{body}'           

            smtp.sendmail(Email_address,EMAIL,msg)
            
            smtp.sendmail(Email_address,email,msg)                
        
            return render(request,"mail.html",succ)
        
